import sqlite3
import hashlib
from datetime import datetime
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

DB_NAME = 'siakternak.db'

DEFAULT_COA = [
    ("101", "Kas & Bank", "Aset"),
    ("102", "Persediaan Sapi Bakalan", "Aset"),
    ("103", "Persediaan Pakan", "Aset"),
    ("104", "Persediaan Obat", "Aset"),
    ("401", "Pendapatan Penjualan Sapi", "Pendapatan"),
    ("501", "Harga Pokok Penjualan Sapi", "HPP"),
    ("502", "Beban Transportasi Pembelian", "HPP"),
    ("601", "Beban Pakan", "Beban"),
    ("602", "Beban Kesehatan Ternak", "Beban"),
    ("603", "Beban Gaji", "Beban"),
    ("604", "Beban Listrik & Air Kandang", "Beban"),
    ("605", "Beban Penyusutan Kandang & Alat", "Beban"),
    ("606", "Beban Operasional Lainnya", "Beban")
]

def get_connection():
    return sqlite3.connect(DB_NAME)

def init_db():
    conn = get_connection()
    c = conn.cursor()

    # Table for users
    c.execute('''CREATE TABLE IF NOT EXISTS users 
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, 
                  username TEXT UNIQUE, password TEXT, 
                  nama_lengkap TEXT, created_at TEXT)''')

    # Table for pemasukan (Income)
    c.execute('''CREATE TABLE IF NOT EXISTS pemasukan 
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, 
                  jumlah_sapi INTEGER, 
                  total_harga INTEGER, 
                  tanggal TEXT)''')

    # Table for pengeluaran (Expense)
    c.execute('''CREATE TABLE IF NOT EXISTS pengeluaran 
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, 
                  produk TEXT, 
                  kategori TEXT, 
                  nominal INTEGER, 
                  jumlah INTEGER DEFAULT 1,
                  tanggal TEXT)''')

    # Table for inventaris (Herd Tracking / Generic Inventory)
    c.execute('''CREATE TABLE IF NOT EXISTS inventaris 
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, 
                  nama_barang TEXT DEFAULT 'Sapi',
                  tipe_transaksi TEXT, 
                  jumlah INTEGER, 
                  tanggal TEXT, 
                  keterangan TEXT,
                  pemasukan_id INTEGER,
                  pengeluaran_id INTEGER,
                  pesanan_id INTEGER)''')

    # Table for pesanan (Buyer Orders)
    c.execute('''CREATE TABLE IF NOT EXISTS pesanan 
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, 
                  nama TEXT, 
                  wa TEXT, 
                  jenis_sapi TEXT, 
                  jumlah_sapi INTEGER, 
                  keterangan TEXT,
                  status TEXT DEFAULT 'Pending',
                  tanggal TEXT)''')

    # Table for COA (Chart of Accounts)
    c.execute('''CREATE TABLE IF NOT EXISTS coa 
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, 
                  code TEXT UNIQUE, 
                  name TEXT, 
                  classification TEXT)''')

    # Dynamic migrations
    c.execute("PRAGMA table_info(pengeluaran)")
    cols_pengeluaran = [col[1] for col in c.fetchall()]
    if 'jumlah' not in cols_pengeluaran:
        c.execute("ALTER TABLE pengeluaran ADD COLUMN jumlah INTEGER DEFAULT 1")

    c.execute("PRAGMA table_info(inventaris)")
    cols_inventaris = [col[1] for col in c.fetchall()]
    if 'nama_barang' not in cols_inventaris:
        c.execute("ALTER TABLE inventaris ADD COLUMN nama_barang TEXT DEFAULT 'Sapi'")
    if 'pesanan_id' not in cols_inventaris:
        c.execute("ALTER TABLE inventaris ADD COLUMN pesanan_id INTEGER")

    # Migrate old transaction types to 'Masuk' / 'Keluar'
    c.execute("UPDATE inventaris SET tipe_transaksi = 'Masuk' WHERE tipe_transaksi IN ('Awal', 'Melahirkan', 'Beli Sapi')")
    c.execute("UPDATE inventaris SET tipe_transaksi = 'Keluar' WHERE tipe_transaksi IN ('Mati', 'Jual Sapi')")

    # Insert default admin if not exists
    c.execute("SELECT * FROM users WHERE username='admin'")
    if not c.fetchone():
        default_password = hashlib.sha256("admin123".encode()).hexdigest()
        c.execute("INSERT INTO users (username, password, nama_lengkap, created_at) VALUES (?,?,?,?)",
                  ("admin", default_password, "Administrator", datetime.now().strftime("%d/%m/%Y %H:%M")))

    # Check if tables are empty to insert dummy data
    c.execute("SELECT COUNT(*) FROM pemasukan")
    has_pemasukan = c.fetchone()[0] > 0
    c.execute("SELECT COUNT(*) FROM pengeluaran")
    has_pengeluaran = c.fetchone()[0] > 0
    c.execute("SELECT COUNT(*) FROM coa")
    has_coa = c.fetchone()[0] > 0
    
    if not has_coa:
        c.executemany("INSERT INTO coa (code, name, classification) VALUES (?,?,?)", DEFAULT_COA)

    if not has_pemasukan and not has_pengeluaran:
        # Insert Pemasukan dummy
        pemasukan_dummies = [
            (5, 75000000, "10/02/2026 10:00"),
            (3, 48000000, "15/03/2026 14:30"),
            (6, 96000000, "02/04/2026 09:15"),
            (4, 64000000, "20/05/2026 16:45"),
            (8, 128000000, "05/06/2026 11:20")
        ]
        c.executemany("INSERT INTO pemasukan (jumlah_sapi, total_harga, tanggal) VALUES (?, ?, ?)", pemasukan_dummies)
        
        # Insert Pengeluaran dummy
        pengeluaran_dummies = [
            ("Pakan Konsentrat", "Pakan", 15000000, 10, "12/02/2026 11:00"),
            ("Vaksin & Obat Sapi", "Kesehatan", 5000000, 4, "18/02/2026 15:30"),
            ("Pakan Hijauan", "Pakan", 12000000, 15, "16/03/2026 10:00"),
            ("Vitamin Organik", "Kesehatan", 4500000, 3, "05/04/2026 14:00"),
            ("Pakan Konsentrat", "Pakan", 18000000, 12, "22/05/2026 11:00"),
            ("Layanan Medis Dokter", "Kesehatan", 8000000, 1, "06/06/2026 12:00")
        ]
        c.executemany("INSERT INTO pengeluaran (produk, kategori, nominal, jumlah, tanggal) VALUES (?, ?, ?, ?, ?)", pengeluaran_dummies)

    # Seed inventaris dummy if empty
    c.execute("SELECT COUNT(*) FROM inventaris")
    if c.fetchone()[0] == 0:
        inventaris_dummies = [
            ("Sapi", "Masuk", 50, "01/01/2026 08:00", "Stok awal tahun", None, None),
            ("Sapi", "Masuk", 5, "10/03/2026 11:00", "Beli bibit unggul", None, None),
            ("Sapi", "Keluar", 1, "25/04/2026 17:00", "Sakit kembung", None, None),
        ]
        c.executemany("INSERT INTO inventaris (nama_barang, tipe_transaksi, jumlah, tanggal, keterangan, pemasukan_id, pengeluaran_id) VALUES (?, ?, ?, ?, ?, ?, ?)",
                      inventaris_dummies)
        
        # Sync the dummy sales into inventaris
        c.execute("SELECT id, jumlah_sapi, tanggal FROM pemasukan")
        sales = c.fetchall()
        for s_id, qty, tgl in sales:
            c.execute("INSERT INTO inventaris (nama_barang, tipe_transaksi, jumlah, tanggal, keterangan, pemasukan_id) VALUES (?, ?, ?, ?, ?, ?)",
                      ("Sapi", "Keluar", qty, tgl, f"Penjualan otomatis (Transaksi ID {s_id})", s_id))

        # Sync the dummy expenses into inventaris
        c.execute("SELECT id, produk, kategori, jumlah, tanggal FROM pengeluaran")
        expenses = c.fetchall()
        for exp_id, prod, cat, qty, tgl in expenses:
            if cat in ["Pakan", "Kesehatan"]:
                item_name = "Pakan Sapi" if cat == "Pakan" else "Obat Sapi"
                c.execute("INSERT INTO inventaris (nama_barang, tipe_transaksi, jumlah, tanggal, keterangan, pengeluaran_id) VALUES (?, ?, ?, ?, ?, ?)",
                          (item_name, "Masuk", qty, tgl, f"Pembelian otomatis (Pengeluaran ID {exp_id})", exp_id))

    conn.commit()
    conn.close()


# --- DATE FILTER UTILITIES ---
def is_in_period(date_str, bulan, tahun):
    if not date_str or len(date_str) < 10:
        return False
    tx_month = date_str[3:5]
    tx_year = date_str[6:10]
    
    match_month = (bulan == "Semua" or tx_month == bulan)
    match_year = (tahun == "Semua" or tx_year == tahun)
    return match_month and match_year

def is_before_period(date_str, bulan, tahun):
    if not date_str or len(date_str) < 10:
        return False
    tx_month = int(date_str[3:5])
    tx_year = int(date_str[6:10])
    
    if tahun != "Semua":
        filter_year = int(tahun)
        if tx_year < filter_year:
            return True
        if tx_year > filter_year:
            return False
            
    if bulan != "Semua":
        filter_month = int(bulan)
        if tx_month < filter_month:
            return True
        if tx_month > filter_month:
            return False
            
    return False


# --- USER MANAGEMENT ---
def verify_user(username, password):
    hashed_password = hashlib.sha256(password.encode()).hexdigest()
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT * FROM users WHERE username=? AND password=?", (username, hashed_password))
    user = c.fetchone()
    conn.close()
    return user

def register_user(username, full_name, password):
    hashed_password = hashlib.sha256(password.encode()).hexdigest()
    tgl = datetime.now().strftime("%d/%m/%Y %H:%M")
    conn = get_connection()
    c = conn.cursor()
    try:
        c.execute("INSERT INTO users (username, password, nama_lengkap, created_at) VALUES (?, ?, ?, ?)",
                  (username, hashed_password, full_name, tgl))
        conn.commit()
        success = True
    except sqlite3.IntegrityError:
        success = False
    conn.close()
    return success

def change_password(username, new_password):
    hashed_password = hashlib.sha256(new_password.encode()).hexdigest()
    conn = get_connection()
    c = conn.cursor()
    c.execute("UPDATE users SET password=? WHERE username=?", (hashed_password, username))
    conn.commit()
    conn.close()

def get_all_users():
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT id, username, nama_lengkap, created_at FROM users ORDER BY id DESC")
    rows = c.fetchall()
    conn.close()
    return rows

def delete_user(username):
    if username == "admin":
        return False
    conn = get_connection()
    c = conn.cursor()
    c.execute("DELETE FROM users WHERE username=?", (username,))
    conn.commit()
    conn.close()
    return True


# --- PEMASUKAN CRUD (WITH AUTO INVENTARIS SYNC) ---
def add_pemasukan(jumlah_sapi, total_harga, tanggal=None):
    if not tanggal:
        tanggal = datetime.now().strftime("%d/%m/%Y %H:%M")
    conn = get_connection()
    c = conn.cursor()
    c.execute("INSERT INTO pemasukan (jumlah_sapi, total_harga, tanggal) VALUES (?, ?, ?)",
              (int(jumlah_sapi), int(total_harga), tanggal))
    pemasukan_id = c.lastrowid
    
    # Sync to inventaris
    c.execute("INSERT INTO inventaris (nama_barang, tipe_transaksi, jumlah, tanggal, keterangan, pemasukan_id) VALUES (?, ?, ?, ?, ?, ?)",
              ("Sapi", "Keluar", int(jumlah_sapi), tanggal, f"Penjualan otomatis (Transaksi ID {pemasukan_id})", pemasukan_id))
    
    conn.commit()
    conn.close()

def get_all_pemasukan():
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT id, jumlah_sapi, total_harga, tanggal FROM pemasukan ORDER BY id DESC")
    rows = c.fetchall()
    conn.close()
    return rows

def update_pemasukan(db_id, jumlah_sapi, total_harga):
    conn = get_connection()
    c = conn.cursor()
    c.execute("UPDATE pemasukan SET jumlah_sapi=?, total_harga=? WHERE id=?",
              (int(jumlah_sapi), int(total_harga), int(db_id)))
    
    # Sync to inventaris
    c.execute("UPDATE inventaris SET jumlah=? WHERE pemasukan_id=?", (int(jumlah_sapi), int(db_id)))
    
    conn.commit()
    conn.close()

def delete_pemasukan(db_id):
    conn = get_connection()
    c = conn.cursor()
    c.execute("DELETE FROM pemasukan WHERE id=?", (int(db_id),))
    
    # Sync to inventaris
    c.execute("DELETE FROM inventaris WHERE pemasukan_id=?", (int(db_id),))
    
    conn.commit()
    conn.close()


# --- PENGELUARAN CRUD (WITH AUTO INVENTARIS SYNC) ---
def add_pengeluaran(produk, kategori, nominal, jumlah=1, tanggal=None):
    if not tanggal:
        tanggal = datetime.now().strftime("%d/%m/%Y %H:%M")
    conn = get_connection()
    c = conn.cursor()
    c.execute("INSERT INTO pengeluaran (produk, kategori, nominal, jumlah, tanggal) VALUES (?, ?, ?, ?, ?)",
              (produk, kategori, int(nominal), int(jumlah), tanggal))
    pengeluaran_id = c.lastrowid
    
    # Sync to inventaris: we sync to inventory using product name
    item_name = produk
    c.execute("INSERT INTO inventaris (nama_barang, tipe_transaksi, jumlah, tanggal, keterangan, pengeluaran_id) VALUES (?, ?, ?, ?, ?, ?)",
              (item_name, "Masuk", int(jumlah), tanggal, f"Pembelian otomatis (Pengeluaran ID {pengeluaran_id})", pengeluaran_id))
    
    conn.commit()
    conn.close()

def get_all_pengeluaran():
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT id, produk, kategori, nominal, jumlah, tanggal FROM pengeluaran ORDER BY id DESC")
    rows = c.fetchall()
    conn.close()
    return rows

def update_pengeluaran(db_id, produk, kategori, nominal, jumlah=1):
    conn = get_connection()
    c = conn.cursor()
    c.execute("UPDATE pengeluaran SET produk=?, kategori=?, nominal=?, jumlah=? WHERE id=?",
              (produk, kategori, int(nominal), int(jumlah), int(db_id)))
    
    # Sync to inventaris
    c.execute("UPDATE inventaris SET nama_barang=?, jumlah=? WHERE pengeluaran_id=?",
              (produk, int(jumlah), int(db_id)))
    
    conn.commit()
    conn.close()

def delete_pengeluaran(db_id):
    conn = get_connection()
    c = conn.cursor()
    c.execute("DELETE FROM pengeluaran WHERE id=?", (int(db_id),))
    
    # Sync to inventaris
    c.execute("DELETE FROM inventaris WHERE pengeluaran_id=?", (int(db_id),))
    
    conn.commit()
    conn.close()


# --- INVENTARIS CRUD (MANUAL EVENTS) ---
def add_inventaris_manual(nama_barang, tipe_transaksi, jumlah, keterangan, tanggal=None):
    if not tanggal:
        tanggal = datetime.now().strftime("%d/%m/%Y %H:%M")
    conn = get_connection()
    c = conn.cursor()
    c.execute("INSERT INTO inventaris (nama_barang, tipe_transaksi, jumlah, tanggal, keterangan) VALUES (?, ?, ?, ?, ?)",
              (nama_barang, tipe_transaksi, int(jumlah), tanggal, keterangan))
    conn.commit()
    conn.close()

def get_all_inventaris():
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT id, nama_barang, tipe_transaksi, jumlah, tanggal, keterangan FROM inventaris ORDER BY id DESC")
    rows = c.fetchall()
    conn.close()
    return rows

def update_inventaris_manual(db_id, nama_barang, tipe_transaksi, jumlah, keterangan):
    conn = get_connection()
    c = conn.cursor()
    c.execute("UPDATE inventaris SET nama_barang=?, tipe_transaksi=?, jumlah=?, keterangan=? WHERE id=?",
              (nama_barang, tipe_transaksi, int(jumlah), keterangan, int(db_id)))
    conn.commit()
    conn.close()

def delete_inventaris_manual(db_id):
    conn = get_connection()
    c = conn.cursor()
    c.execute("DELETE FROM inventaris WHERE id=?", (int(db_id),))
    conn.commit()
    conn.close()

def get_inventory_summary():
    conn = get_connection()
    c = conn.cursor()
    
    # Calculate stock for Sapi
    c.execute("SELECT SUM(jumlah) FROM inventaris WHERE nama_barang='Sapi' AND tipe_transaksi='Masuk'")
    sapi_in = c.fetchone()[0] or 0
    c.execute("SELECT SUM(jumlah) FROM inventaris WHERE nama_barang='Sapi' AND tipe_transaksi='Keluar'")
    sapi_out = c.fetchone()[0] or 0
    sapi_stock = sapi_in - sapi_out
    
    # Calculate stock for Pakan
    c.execute("SELECT SUM(jumlah) FROM inventaris WHERE nama_barang LIKE '%Pakan%' AND tipe_transaksi='Masuk'")
    pakan_in = c.fetchone()[0] or 0
    c.execute("SELECT SUM(jumlah) FROM inventaris WHERE nama_barang LIKE '%Pakan%' AND tipe_transaksi='Keluar'")
    pakan_out = c.fetchone()[0] or 0
    pakan_stock = pakan_in - pakan_out
    
    # Calculate stock for Obat
    c.execute("SELECT SUM(jumlah) FROM inventaris WHERE (nama_barang LIKE '%Obat%' OR nama_barang LIKE '%Vaksin%' OR nama_barang LIKE '%Kesehatan%') AND tipe_transaksi='Masuk'")
    obat_in = c.fetchone()[0] or 0
    c.execute("SELECT SUM(jumlah) FROM inventaris WHERE (nama_barang LIKE '%Obat%' OR nama_barang LIKE '%Vaksin%' OR nama_barang LIKE '%Kesehatan%') AND tipe_transaksi='Keluar'")
    obat_out = c.fetchone()[0] or 0
    obat_stock = obat_in - obat_out
    
    conn.close()
    
    return {
        'sapi': sapi_stock,
        'pakan': pakan_stock,
        'obat': obat_stock
    }


# --- SUMMARY & DASHBOARD ---
def get_summary():
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT SUM(total_harga) FROM pemasukan")
    total_in = c.fetchone()[0] or 0
    c.execute("SELECT SUM(nominal) FROM pengeluaran")
    total_out = c.fetchone()[0] or 0
    conn.close()
    return {
        'total_in': total_in,
        'total_out': total_out,
        'laba': total_in - total_out
    }

def get_recent_history(limit=20):
    conn = get_connection()
    c = conn.cursor()
    query = """
    SELECT tipe, id, detail, nominal, tanggal FROM (
        SELECT 'Pemasukan' as tipe, id, jumlah_sapi || ' Sapi' as detail, total_harga as nominal, tanggal 
        FROM pemasukan
        UNION ALL
        SELECT 'Pengeluaran' as tipe, id, produk || ' (' || kategori || ')' as detail, nominal, tanggal
        FROM pengeluaran
    )
    ORDER BY substr(tanggal, 7, 4) DESC, substr(tanggal, 4, 2) DESC, substr(tanggal, 1, 2) DESC, substr(tanggal, 12, 5) DESC
    LIMIT ?
    """
    c.execute(query, (limit,))
    rows = c.fetchall()
    conn.close()
    return rows

def get_monthly_summary(limit=6):
    conn = get_connection()
    c = conn.cursor()
    pemasukan_query = """
    SELECT substr(tanggal, 7, 4) || '-' || substr(tanggal, 4, 2) as yyyymm, SUM(total_harga)
    FROM pemasukan
    GROUP BY yyyymm
    """
    c.execute(pemasukan_query)
    in_data = dict(c.fetchall())
    
    pengeluaran_query = """
    SELECT substr(tanggal, 7, 4) || '-' || substr(tanggal, 4, 2) as yyyymm, SUM(nominal)
    FROM pengeluaran
    GROUP BY yyyymm
    """
    c.execute(pengeluaran_query)
    out_data = dict(c.fetchall())
    conn.close()
    
    all_months = sorted(list(set(in_data.keys()) | set(out_data.keys())))
    if not all_months:
        current_m = datetime.now().strftime("%Y-%m")
        all_months = [current_m]
        
    all_months = all_months[-limit:]
    
    month_names = {
        "01": "Jan", "02": "Feb", "03": "Mar", "04": "Apr", "05": "Mei", "06": "Jun",
        "07": "Jul", "08": "Agt", "09": "Sep", "10": "Okt", "11": "Nov", "12": "Des"
    }
    
    labels = []
    pemasukan_vals = []
    pengeluaran_vals = []
    
    for ym in all_months:
        if '-' in ym:
            year, month = ym.split('-')
            lbl = f"{month_names.get(month, month)} {year[2:]}"
        else:
            lbl = ym
        labels.append(lbl)
        pemasukan_vals.append(in_data.get(ym, 0))
        pengeluaran_vals.append(out_data.get(ym, 0))
        
    return labels, pemasukan_vals, pengeluaran_vals


# --- EXCEL EXPORT ---
def export_to_excel(filepath="Laporan_Siakternak.xlsx"):
    wb = openpyxl.Workbook()
    
    # Sheet 1: Pemasukan
    ws_in = wb.active
    ws_in.title = "Pemasukan"
    headers_in = ["No", "Jumlah Sapi (Ekor)", "Total Penjualan (Rp)", "Tanggal"]
    ws_in.append(headers_in)
    rows_in = get_all_pemasukan()
    for i, r in enumerate(reversed(rows_in), start=1):
        ws_in.append([i, r[1], r[2], r[3]])
    style_sheet(ws_in, "A1:D1", ["#FFD6D6D6", "#FF00FF00"])
    
    # Sheet 2: Pengeluaran
    ws_out = wb.create_sheet(title="Pengeluaran")
    headers_out = ["No", "Produk", "Kategori", "Nominal (Rp)", "Jumlah", "Tanggal"]
    ws_out.append(headers_out)
    rows_out = get_all_pengeluaran()
    for i, r in enumerate(reversed(rows_out), start=1):
        # r = (id, produk, kategori, nominal, jumlah, tanggal)
        ws_out.append([i, r[1], r[2], r[3], r[4], r[5]])
    style_sheet(ws_out, "A1:F1", ["#FFD6D6D6", "#FFFF0000"])
    
    # Sheet 3: Inventaris Ternak
    ws_inv = wb.create_sheet(title="Inventaris Ternak")
    headers_inv = ["No", "Nama Barang / Hal", "Tipe Mutasi", "Jumlah", "Tanggal", "Keterangan"]
    ws_inv.append(headers_inv)
    rows_inv = get_all_inventaris()
    for i, r in enumerate(reversed(rows_inv), start=1):
        # r = (id, nama_barang, tipe_transaksi, jumlah, tanggal, keterangan)
        ws_inv.append([i, r[1], r[2], r[3], r[4], r[5]])
    style_sheet(ws_inv, "A1:F1", ["#FFD6D6D6", "#FF0000FF"])
    
    wb.save(filepath)
    return os.path.abspath(filepath)

def style_sheet(ws, header_range, colors):
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    if "00FF00" in colors[1]:
        fill_color = "1B5E20"
    elif "FF0000" in colors[1]:
        fill_color = "B71C1C"
    else:
        fill_color = "1A237E" # dark blue for inventory
        
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            if isinstance(cell.value, int):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")
            elif cell.column == 1:
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left")
                
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)


# --- ACCOUNTING MODULE HELPERS (WITH PERIOD FILTERS) ---
def get_coa():
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT code, name, classification FROM coa ORDER BY code")
    rows = c.fetchall()
    conn.close()
    return rows


def add_coa_account(code, name, classification):
    conn = get_connection()
    c = conn.cursor()
    try:
        c.execute("INSERT INTO coa (code, name, classification) VALUES (?,?,?)", (code, name, classification))
        conn.commit()
        return True
    except sqlite3.IntegrityError:
        return False
    finally:
        conn.close()


def map_expense_account(produk, kategori):
    cat_lower = kategori.lower()
    prod_lower = produk.lower()
    
    if "pakan" in cat_lower or "pakan" in prod_lower:
        return "103", "Persediaan Pakan"
    elif "kesehatan" in cat_lower or "obat" in cat_lower or "vaksin" in prod_lower:
        return "104", "Persediaan Obat"
    elif "gaji" in cat_lower or "gaji" in prod_lower:
        return "603", "Beban Gaji"
    elif "listrik" in cat_lower or "air" in cat_lower or "listrik" in prod_lower or "air" in prod_lower:
        return "604", "Beban Listrik & Air Kandang"
    elif "penyusutan" in cat_lower or "penyusutan" in prod_lower:
        return "605", "Beban Penyusutan Kandang & Alat"
    elif "transport" in cat_lower or "angkut" in cat_lower or "transport" in prod_lower or "angkut" in prod_lower:
        return "502", "Beban Transportasi Pembelian"
    elif "sapi" in cat_lower or "bakalan" in cat_lower or "sapi" in prod_lower or "bakalan" in prod_lower:
        return "102", "Persediaan Sapi Bakalan"
    else:
        return "606", "Beban Operasional Lainnya"

def get_jurnal_umum(bulan="Semua", tahun="Semua"):
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT jumlah_sapi, total_harga, tanggal FROM pemasukan")
    pemasukan_rows = c.fetchall()
    c.execute("SELECT produk, kategori, nominal, tanggal FROM pengeluaran")
    pengeluaran_rows = c.fetchall()
    conn.close()
    
    all_tx = []
    for row in pemasukan_rows:
        all_tx.append({
            'tipe': 'Pemasukan',
            'val': row[1],
            'tanggal': row[2],
            'desc': f"Penjualan {row[0]} Sapi"
        })
    for row in pengeluaran_rows:
        all_tx.append({
            'tipe': 'Pengeluaran',
            'val': row[2],
            'produk': row[0],
            'kategori': row[1],
            'tanggal': row[3],
            'desc': f"Pembelian {row[0]}"
        })
        
    def parse_date(date_str):
        try:
            return datetime.strptime(date_str, "%d/%m/%Y %H:%M")
        except:
            return datetime.min
            
    all_tx.sort(key=lambda x: parse_date(x['tanggal']))
    
    journal_rows = []
    for tx in all_tx:
        tgl = tx['tanggal']
        if not is_in_period(tgl, bulan, tahun):
            continue
            
        if tx['tipe'] == 'Pemasukan':
            val = tx['val']
            journal_rows.append((tgl, "101", "Kas & Bank", f"Rp {val:,}", ""))
            journal_rows.append((tgl, "401", "  Pendapatan Penjualan Sapi", "", f"Rp {val:,}"))
            
            # Debit HPP (501) & Credit Persediaan Sapi Bakalan (102)
            hpp_val = int(val * 0.7)
            journal_rows.append((tgl, "501", "Harga Pokok Penjualan Sapi", f"Rp {hpp_val:,}", ""))
            journal_rows.append((tgl, "102", "  Persediaan Sapi Bakalan", "", f"Rp {hpp_val:,}"))
        else:
            val = tx['val']
            code, name = map_expense_account(tx['produk'], tx['kategori'])
            
            # Debit asset/expense & Credit Kas (101)
            journal_rows.append((tgl, code, name, f"Rp {val:,}", ""))
            journal_rows.append((tgl, "101", "  Kas & Bank", "", f"Rp {val:,}"))
            
            # Accrual usage adjustments for Pakan and Obat
            if code == "103":
                use_val = int(val * 0.8)
                journal_rows.append((tgl, "601", "Beban Pakan", f"Rp {use_val:,}", ""))
                journal_rows.append((tgl, "103", "  Persediaan Pakan", "", f"Rp {use_val:,}"))
            elif code == "104":
                use_val = int(val * 0.8)
                journal_rows.append((tgl, "602", "Beban Kesehatan Ternak", f"Rp {use_val:,}", ""))
                journal_rows.append((tgl, "104", "  Persediaan Obat", "", f"Rp {use_val:,}"))
            
    return journal_rows

def get_buku_besar(kode_akun, bulan="Semua", tahun="Semua"):
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT jumlah_sapi, total_harga, tanggal FROM pemasukan")
    pemasukan_rows = c.fetchall()
    c.execute("SELECT produk, kategori, nominal, tanggal FROM pengeluaran")
    pengeluaran_rows = c.fetchall()
    conn.close()
    
    all_tx = []
    for row in pemasukan_rows:
        all_tx.append({
            'tipe': 'Pemasukan',
            'val': row[1],
            'tanggal': row[2],
            'desc': f"Penjualan {row[0]} Sapi"
        })
    for row in pengeluaran_rows:
        all_tx.append({
            'tipe': 'Pengeluaran',
            'val': row[2],
            'produk': row[0],
            'kategori': row[1],
            'tanggal': row[3],
            'desc': f"Pembelian {row[0]}"
        })
        
    def parse_date(date_str):
        try:
            return datetime.strptime(date_str, "%d/%m/%Y %H:%M")
        except:
            return datetime.min
            
    all_tx.sort(key=lambda x: parse_date(x['tanggal']))
    
    ledger_rows = []
    beginning_balance = 0
    has_filter = (bulan != "Semua" or tahun != "Semua")
    
    # Beginning balances for assets
    if kode_akun == "101":
        beginning_balance = 500_000_000
    elif kode_akun == "102":
        beginning_balance = 150_000_000
    elif kode_akun == "103":
        beginning_balance = 20_000_000
    elif kode_akun == "104":
        beginning_balance = 10_000_000
        
    balance = beginning_balance
    
    for tx in all_tx:
        tgl = tx['tanggal']
        desc = tx['desc']
        debit = 0
        credit = 0
        affected = False
        
        if tx['tipe'] == 'Pemasukan':
            val = tx['val']
            if kode_akun == "101":
                debit = val
                affected = True
            elif kode_akun == "401":
                credit = val
                affected = True
            elif kode_akun == "501":
                debit = int(val * 0.7)
                affected = True
            elif kode_akun == "102":
                credit = int(val * 0.7)
                affected = True
        else:
            val = tx['val']
            code_dest, _ = map_expense_account(tx['produk'], tx['kategori'])
            if kode_akun == "101":
                credit = val
                affected = True
            elif kode_akun == code_dest:
                debit = val
                affected = True
            
            # Pakan usage adjustment
            if code_dest == "103":
                use_val = int(val * 0.8)
                if kode_akun == "601":
                    debit = use_val
                    affected = True
                elif kode_akun == "103":
                    credit = use_val
                    affected = True
            
            # Obat usage adjustment
            if code_dest == "104":
                use_val = int(val * 0.8)
                if kode_akun == "602":
                    debit = use_val
                    affected = True
                elif kode_akun == "104":
                    credit = use_val
                    affected = True
                
        if affected:
            coa = get_coa()
            classification = next((item[2] for item in coa if item[0] == kode_akun), "Beban")
            is_debit_normal = (classification in ["Aset", "HPP", "Beban"])
            change = (debit - credit) if is_debit_normal else (credit - debit)
            
            if has_filter and is_before_period(tgl, bulan, tahun):
                beginning_balance += change
            elif is_in_period(tgl, bulan, tahun) or not has_filter:
                if len(ledger_rows) == 0 and has_filter:
                    ledger_rows.append((" - ", "Saldo Awal / Pindahan", "", "", f"Rp {beginning_balance:,}"))
                    balance = beginning_balance
                
                balance += change
                debit_str = f"Rp {debit:,}" if debit > 0 else ""
                credit_str = f"Rp {credit:,}" if credit > 0 else ""
                ledger_rows.append((tgl, desc, debit_str, credit_str, f"Rp {balance:,}"))
                
    if has_filter and len(ledger_rows) == 0:
        ledger_rows.append((" - ", "Saldo Awal / Pindahan", "", "", f"Rp {beginning_balance:,}"))
        
    return ledger_rows

def get_neraca_saldo(bulan="Semua", tahun="Semua"):
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT jumlah_sapi, total_harga, tanggal FROM pemasukan")
    pemasukan_rows = c.fetchall()
    c.execute("SELECT produk, kategori, nominal, tanggal FROM pengeluaran")
    pengeluaran_rows = c.fetchall()
    conn.close()
    
    balances = {
        "101": 500_000_000,
        "102": 150_000_000,
        "103": 20_000_000,
        "104": 10_000_000,
        "401": 0,
        "501": 0,
        "502": 0,
        "601": 0,
        "602": 0,
        "603": 0,
        "604": 0,
        "605": 0,
        "606": 0
    }
    
    for row in pemasukan_rows:
        qty, val, tgl = row
        if is_before_period(tgl, bulan, tahun) or is_in_period(tgl, bulan, tahun):
            balances["101"] += val
            hpp_val = int(val * 0.7)
            balances["102"] -= hpp_val
            
            if is_in_period(tgl, bulan, tahun):
                balances["401"] += val
                balances["501"] += hpp_val
            elif not is_before_period(tgl, bulan, tahun) and not is_in_period(tgl, bulan, tahun):
                pass
            else:
                # Cumulative assets before/in period must include HPP reductions
                pass
                
    for row in pengeluaran_rows:
        prod, cat, val, tgl = row
        if is_before_period(tgl, bulan, tahun) or is_in_period(tgl, bulan, tahun):
            balances["101"] -= val
            code, _ = map_expense_account(prod, cat)
            
            if is_in_period(tgl, bulan, tahun):
                if code == "103":
                    balances["103"] += val
                    balances["103"] -= int(val * 0.8)
                    balances["601"] += int(val * 0.8)
                elif code == "104":
                    balances["104"] += val
                    balances["104"] -= int(val * 0.8)
                    balances["602"] += int(val * 0.8)
                else:
                    balances[code] += val
            elif is_before_period(tgl, bulan, tahun):
                # Apply changes to cumulative balance sheet assets
                if code == "103":
                    balances["103"] += val - int(val * 0.8)
                elif code == "104":
                    balances["104"] += val - int(val * 0.8)
                elif code in ["102"]:
                    balances[code] += val
                # Expense accounts don't carry over as they are closed, but we keep calculations clean
                
    rows = []
    total_debit = 0
    total_kredit = 0
    coa = get_coa()
    
    for code, name, classification in coa:
        bal = balances[code]
        is_debit_normal = (classification in ["Aset", "HPP", "Beban"])
        
        debit_str = ""
        kredit_str = ""
        
        if is_debit_normal:
            if bal >= 0:
                debit_str = f"Rp {bal:,}"
                total_debit += bal
            else:
                kredit_str = f"Rp {abs(bal):,}"
                total_kredit += abs(bal)
        else:
            if bal >= 0:
                kredit_str = f"Rp {bal:,}"
                total_kredit += bal
            else:
                debit_str = f"Rp {abs(bal):,}"
                total_debit += abs(bal)
                
        rows.append((code, name, debit_str, kredit_str))
        
    return rows, total_debit, total_kredit

def get_laba_rugi(bulan="Semua", tahun="Semua"):
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT total_harga, tanggal FROM pemasukan")
    pemasukan_rows = c.fetchall()
    c.execute("SELECT produk, kategori, nominal, tanggal FROM pengeluaran")
    pengeluaran_rows = c.fetchall()
    conn.close()
    
    pb = {
        "401": 0, "501": 0, "502": 0,
        "601": 0, "602": 0, "603": 0, "604": 0, "605": 0, "606": 0
    }
    
    for row in pemasukan_rows:
        val, tgl = row
        if is_in_period(tgl, bulan, tahun):
            pb["401"] += val
            pb["501"] += int(val * 0.7)
            
    for row in pengeluaran_rows:
        prod, cat, val, tgl = row
        if is_in_period(tgl, bulan, tahun):
            code, _ = map_expense_account(prod, cat)
            if code == "103":
                pb["601"] += int(val * 0.8)
            elif code == "104":
                pb["602"] += int(val * 0.8)
            elif code in pb:
                pb[code] += val
                
    total_revenue = pb["401"]
    total_hpp = pb["501"] + pb["502"]
    total_beban = pb["601"] + pb["602"] + pb["603"] + pb["604"] + pb["605"] + pb["606"]
    laba_bersih = total_revenue - total_hpp - total_beban
    
    return {
        'pendapatan': total_revenue,
        'hpp_sapi': pb["501"],
        'beban_transport': pb["502"],
        'total_hpp': total_hpp,
        'beban_pakan': pb["601"],
        'beban_kesehatan': pb["602"],
        'beban_gaji': pb["603"],
        'beban_listrik': pb["604"],
        'beban_penyusutan': pb["605"],
        'beban_lain': pb["606"],
        'total_beban': total_beban,
        'laba_bersih': laba_bersih
    }

# --- BUYER ORDERS (PESANAN) ---
def add_pesanan(nama, wa, jenis_sapi, jumlah_sapi, keterangan, tanggal=None):
    if not tanggal:
        tanggal = datetime.now().strftime("%d/%m/%Y %H:%M")
    conn = get_connection()
    c = conn.cursor()
    c.execute("INSERT INTO pesanan (nama, wa, jenis_sapi, jumlah_sapi, keterangan, status, tanggal) VALUES (?, ?, ?, ?, ?, ?, ?)",
              (nama, wa, jenis_sapi, int(jumlah_sapi), keterangan, 'Pending', tanggal))
    conn.commit()
    conn.close()

def get_all_pesanan():
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT id, nama, wa, jenis_sapi, jumlah_sapi, keterangan, status, tanggal FROM pesanan ORDER BY id DESC")
    rows = c.fetchall()
    conn.close()
    return rows

def update_status_pesanan(pesanan_id, status):
    conn = get_connection()
    c = conn.cursor()
    
    # Get order info
    c.execute("SELECT nama, jumlah_sapi, tanggal FROM pesanan WHERE id = ?", (int(pesanan_id),))
    row = c.fetchone()
    if not row:
        conn.close()
        return False
    
    nama, jumlah_sapi, tanggal = row
    
    # Update status
    c.execute("UPDATE pesanan SET status = ? WHERE id = ?", (status, int(pesanan_id)))
    
    if status == 'Iya':
        # Check if already synced in inventaris
        c.execute("SELECT id FROM inventaris WHERE pesanan_id = ?", (int(pesanan_id),))
        inv_row = c.fetchone()
        if not inv_row:
            # Sync to inventaris (transaksi Keluar)
            c.execute("INSERT INTO inventaris (nama_barang, tipe_transaksi, jumlah, tanggal, keterangan, pesanan_id) VALUES (?, ?, ?, ?, ?, ?)",
                      ("Sapi", "Keluar", int(jumlah_sapi), tanggal, f"Acc Pemesanan ID {pesanan_id} ({nama})", int(pesanan_id)))
    else:
        # If status is set to 'Tidak' or 'Pending', remove synced row if any
        c.execute("DELETE FROM inventaris WHERE pesanan_id = ?", (int(pesanan_id),))
        
    conn.commit()
    conn.close()
    return True
